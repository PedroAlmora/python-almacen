from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseNotFound, JsonResponse, HttpResponse, FileResponse
from django.contrib import messages
from .models import Producto, Bulto, Almacen, Estante, Despacho
from .forms import ProductoForm, EscanearCodigoDeBarrasForm, AlmacenForm, EstanteForm, EscanearCodigoDeBarrasDespachoForm, CargarProductosForm, DespachoForm
from django.db.models import Sum
import logging
from django.db.models import Q
from .util import *
from django.http import Http404
import logging
#################################
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from datetime import date


###################LOGS##################
@login_required(login_url='login_app:login')
def LoggerUser(request):
    logs = obtenerActividadesLogger()
    return render(request, "logs_user.html", 
                  { "logs" : logs})

logger = logging.getLogger(__name__)


def server_error(request):
    error_type = "Error del servidor"
    error_message = "Lo sentimos, se ha producido un error interno en el servidor."
    return render(request, 'error.html', {'error_type': error_type, 'error_message': error_message})

def delete_logger(request, logger_id):
    logger_entry = get_object_or_404(LoggerUsuario, id=logger_id)
    logger_entry.delete()
    add_bootstrap_message(request, 'success', '¡Logger eliminado exitosamente!')
    return redirect('almacen_app:logsUsers')
#############################TODOS##################
@login_required(login_url='login_app:login')
def Productos(request):
    productos = Producto.objects.all().order_by('codigo')
    return render(request, "productos.html", 
                  { "productos" : productos})
    
@login_required(login_url='login_app:login')
def agotados(request):
    productos_agotados = Producto.objects.filter(cantidad= 0).order_by('codigo')
    return render(request, "categorias-productos/agotados.html", { "agotados": productos_agotados })
#############################ACCESORIOS##################
@login_required(login_url='login_app:login')
def neumaticos(request):
    productos_neumaticos = Producto.objects.filter(campo_predefinido='Neumáticos y productos relacionados').order_by('codigo')
    return render(request, "categorias-productos/neumaticos.html", { "neumaticos": productos_neumaticos })

@login_required(login_url='login_app:login')
def frenos(request):
    productos_frenos = Producto.objects.filter(campo_predefinido='Frenos').order_by('codigo')
    return render(request, "categorias-productos/frenos.html", { "frenos": productos_frenos })

@login_required(login_url='login_app:login')
def filtros(request):
    productos_filtros = Producto.objects.filter(campo_predefinido='Filtros').order_by('codigo')
    return render(request, "categorias-productos/filtros.html", { "filtros": productos_filtros })

@login_required(login_url='login_app:login')
def motor(request):
    productos_motor = Producto.objects.filter(campo_predefinido='Motor').order_by('codigo')
    return render(request, "categorias-productos/motor.html", { "motores": productos_motor })

@login_required(login_url='login_app:login')
def aceites_y_lubricantes(request):
    productos_aceites_lubricantes = Producto.objects.filter(campo_predefinido='Aceites y líquidos').order_by('codigo')
    return render(request, "categorias-productos/aceites_lubricantes.html", { "aceites_lubricantes": productos_aceites_lubricantes })

@login_required(login_url='login_app:login')
def limpiaparabrisas_y_escobillas(request):
    productos_limpiaparabrisas_escobillas = Producto.objects.filter(campo_predefinido='Sistema limpiaparabrisas').order_by('codigo')
    return render(request, "categorias-productos/limpiaparabrisas_escobillas.html", { "limpiaparabrisas_escobillas": productos_limpiaparabrisas_escobillas })

@login_required(login_url='login_app:login')
def suspension(request):
    productos_suspension = Producto.objects.filter(campo_predefinido='Suspensión').order_by('codigo')
    return render(request, "categorias-productos/suspension.html", { "suspension": productos_suspension })

@login_required(login_url='login_app:login')
def sistema_electrico(request):
    productos_sistema_electrico = Producto.objects.filter(campo_predefinido='Sistema eléctrico').order_by('codigo')
    return render(request, "categorias-productos/sistema_electrico.html", { "sistema_electrico": productos_sistema_electrico })

@login_required(login_url='login_app:login')
def endendido_y_precalentamiento(request):
    productos_encendido_precalentamiento = Producto.objects.filter(campo_predefinido='Encendido y precalentamiento').order_by('codigo')
    return render(request, "categorias-productos/encendido_precalentamiento.html", { "encendido_precalentamiento": productos_encendido_precalentamiento })

@login_required(login_url='login_app:login')
def correas_cadenas_rodillos(request):
    productos_correas_cadenas_rodillos = Producto.objects.filter(campo_predefinido='Correas, cadenas, rodillos').order_by('codigo')
    return render(request, "categorias-productos/correas_cadenas_rodillos.html", { "correas_cadenas_rodillos": productos_correas_cadenas_rodillos })

@login_required(login_url='login_app:login')
def carroceria(request):
    productos_carroceria = Producto.objects.filter(campo_predefinido='Carrocería').order_by('codigo')
    return render(request, "categorias-productos/carroceria.html", { "carrocerias": productos_carroceria })

@login_required(login_url='login_app:login')
def calefaccion_y_ventilacion(request):
    productos_calefaccion_ventilacion = Producto.objects.filter(campo_predefinido='Calefacción y ventilación').order_by('codigo')
    return render(request, "categorias-productos/calefaccion_ventilacion.html", { "calefaccion_ventilacion": productos_calefaccion_ventilacion })

@login_required(login_url='login_app:login')
def baterias(request):
    productos_baterias = Producto.objects.filter(campo_predefinido='Escape').order_by('codigo')
    return render(request, "categorias-productos/baterias.html", { "baterias": productos_baterias })

@login_required(login_url='login_app:login')
def sistema_de_combustible(request):
    productos_sistema_combustible = Producto.objects.filter(campo_predefinido='Sistema de combustible').order_by('codigo')
    return render(request, "categorias-productos/sistema_combustible.html", { "sistema_combustible": productos_sistema_combustible })

@login_required(login_url='login_app:login')
def direccion(request):
    productos_direccion = Producto.objects.filter(campo_predefinido='Dirección').order_by('codigo')
    return render(request, "categorias-productos/direccion.html", { "direccion": productos_direccion })

@login_required(login_url='login_app:login')
def amortiguacion(request):
    productos_amortiguacion = Producto.objects.filter(campo_predefinido='Amortiguación').order_by('codigo')
    return render(request, "categorias-productos/amortiguacion.html", { "amortiguacion": productos_amortiguacion })

@login_required(login_url='login_app:login')
def sistema_de_refrigeracion(request):
    productos_sistema_refrigeracion = Producto.objects.filter(campo_predefinido='Sistema de refrigeración del motor').order_by('codigo')
    return render(request, "categorias-productos/sistema_refrigeracion.html", { "sistema_refrigeracion": productos_sistema_refrigeracion })

@login_required(login_url='login_app:login')
def juntas_y_retrenes(request):
    productos_juntas_retretes = Producto.objects.filter(campo_predefinido='Juntas y retenes').order_by('codigo')
    return render(request, "categorias-productos/juntas_retretes.html", { "juntas_retrenes": productos_juntas_retretes })

@login_required(login_url='login_app:login')
def interior(request):
    productos_interior = Producto.objects.filter(campo_predefinido='Interior').order_by('codigo')
    return render(request, "categorias-productos/interior.html", { "interior": productos_interior })

@login_required(login_url='login_app:login')
def embrague(request):
    productos_embrague = Producto.objects.filter(campo_predefinido='Embrague').order_by('codigo')
    return render(request, "categorias-productos/embrague.html", { "embrague": productos_embrague })

@login_required(login_url='login_app:login')
def caja_de_cambios(request):
    productos_caja_cambios = Producto.objects.filter(campo_predefinido='Caja de cambios').order_by('codigo')
    return render(request, "categorias-productos/caja_cambios.html", { "caja_cambios": productos_caja_cambios })

@login_required(login_url='login_app:login')
def aire_acondicionado(request):
    productos_aire_acondicionado = Producto.objects.filter(campo_predefinido='Aire acondicionado').order_by('codigo')
    return render(request, "categorias-productos/aire_acondicionado.html", { "aire_acondicionado": productos_aire_acondicionado })

@login_required(login_url='login_app:login')
def rodamientos(request):
    productos_rodamientos = Producto.objects.filter(campo_predefinido='Rodamientos').order_by('codigo')
    return render(request, "categorias-productos/rodamientos.html", { "rodamientos": productos_rodamientos })

@login_required(login_url='login_app:login')
def transmision_y_diferencial(request):
    productos_transmision_diferencial = Producto.objects.filter(campo_predefinido='Árboles de transmisión y diferenciales').order_by('codigo')
    return render(request, "categorias-productos/transmision_diferencial.html", { "transmision_diferencial": productos_transmision_diferencial })

@login_required(login_url='login_app:login')
def kit_de_reparacion(request):
    productos_kit_reparacion = Producto.objects.filter(campo_predefinido='Kits de reparación Y Herramientas').order_by('codigo')
    return render(request, "categorias-productos/kit_reparacion.html", { "kit_reparacion": productos_kit_reparacion })

@login_required(login_url='login_app:login')
def accesorios(request):
    productos_accesorios = Producto.objects.filter(campo_predefinido='Accesorios para coches').order_by('codigo')
    return render(request, "categorias-productos/accesorios.html", { "accesorios": productos_accesorios })

@login_required(login_url='login_app:login')
def tuberias_y_mangueras(request):
    productos_tuberias_mangueras = Producto.objects.filter(campo_predefinido='Tuberías y mangueras').order_by('codigo')
    return render(request, "categorias-productos/tuberias_mangueras.html", { "tuberias_mangueras": productos_tuberias_mangueras })

@login_required(login_url='login_app:login')
def pailer_juntas_homocinetica(request):
    pailer_juntas_homocineticas = Producto.objects.filter(campo_predefinido='Palier y junta homocinética').order_by('codigo')
    return render(request, "categorias-productos/pailer_juntas_homocineticas.html", { "pailer_juntas_homocineticas": pailer_juntas_homocineticas })

@login_required(login_url='login_app:login')
def remolque_piezas_adicionales(request):
    remolque_piezas_adicionales = Producto.objects.filter(campo_predefinido='Remolque / piezas adicionales').order_by('codigo')
    return render(request, "categorias-productos/remolque_piezas_adicionales.html", { "remolque_piezas_adicionales": remolque_piezas_adicionales })

@login_required(login_url='login_app:login')
def sensores_reles(request):
    sensores_reles = Producto.objects.filter(campo_predefinido='Sensores, relés, unidades de control').order_by('codigo')
    return render(request, "categorias-productos/sensores_reles.html", { "sensores_reles": sensores_reles })

@login_required(login_url='login_app:login')
def iluminacion(request):
    iluminacion = Producto.objects.filter(campo_predefinido='Iluminación').order_by('codigo')
    return render(request, "categorias-productos/iluminacion.html", { "iluminacion": iluminacion })

@login_required(login_url='login_app:login')
def sujeciones(request):
    sujeciones = Producto.objects.filter(campo_predefinido='Sujeciones').order_by('codigo')
    return render(request, "categorias-productos/sujeciones.html", { "sujeciones": sujeciones })
######################FORMULARIO-PRODUCTO###############################

######################CREATE-PRODUCTO###################################

@login_required(login_url='login_app:login')
def create_product(request):
    try:
        if request.method == 'POST':
            form = ProductoForm(request.POST, request.FILES)
            if form.is_valid():
                producto = form.save(commit=False)  
                producto.save()
                registrarLoggerUsuario(request.user, f"El usuario insertó el producto '{producto.nombre}' en la base de datos")
                add_bootstrap_message(request, 'success', '¡Producto creado exitosamente!')
                return redirect('almacen_app:create_producto')
                add_bootstrap_message(request, 'warning', 'Por favor. Verifique que los campos estén correctos')
        else:
            form = ProductoForm()

        context = {
            'form': form,
            'is_editing': False,
        }
        return render(request, 'product_forms.html', context)
    except Exception:
        return server_error(request) 

#############################################################################

##########################EDIT-PRODUCTO######################################

@login_required(login_url='login_app:login')
def edit_product(request, product_id):
    try:
        product = Producto.objects.get(id=product_id)
        if request.method == 'POST':
            form = ProductoForm(request.POST, request.FILES, instance=product)
            if form.is_valid():
                edited_product = form.save(commit=False)
                edited_fields = get_edited_fields(product, edited_product)
                if edited_fields:
                    registrarLoggerUsuario(request.user, f"El usuario editó el producto '{product.nombre}', campos editados: '{edited_fields}'")
                    edited_product.save()
                else:
                    registrarLoggerUsuario(request.user, f"El usuario editó el producto '{product.nombre}' en la base de datos")
                add_bootstrap_message(request, 'success', '¡Producto actualizado exitosamente!')
                edited_product.save()
                return redirect('almacen_app:productos')
            else:
                add_bootstrap_message(request, 'warning', 'Por favor. Verifique que los campos estén correctos')
        else:
            form = ProductoForm(instance=product)

        context = {
            'form': form,
            'is_editing': True,
        }
        return render(request, 'product_forms.html', context)  
    except Producto.DoesNotExist:
        return HttpResponseNotFound("Producto no encontrado")
    except Exception:
        return server_error(request)

def get_edited_fields(old_product, new_product):
    edited_fields = []
    fields = ['codigo', 'codigo_barras', 'nombre', 'descripcion', 'precio', 'cantidad', 'compatibilidad', 'imagen', 'campo_predefinido']  
    for field in fields:
        if getattr(old_product, field) != getattr(new_product, field):
            edited_fields.append(field)
    return edited_fields
###########################################################################
@login_required(login_url='login_app:login')
def delete_product(request, product_id):
    product = get_object_or_404(Producto, id=product_id)
    try:
        if request.method == 'POST':
            product_name = product.nombre 
            product.delete()
            registrarLoggerUsuario(request.user, f"El usuario eliminó el producto '{product_name}' en la base de datos")
            add_bootstrap_message(request, 'success', '¡Producto eliminado exitosamente!')
            return redirect('almacen_app:productos')

        context = {
            'product': product,
        }
        return render(request, 'confirm_delete.html', context)
    except Exception:
        return server_error(request)

###########################################################################

########################LECTOR-CODIGO-BARRAS###############################
@login_required(login_url='login_app:login')
def escanear_codigo_barras(request):
    try:
        form = EscanearCodigoDeBarrasForm(request.POST or None)
        bulto = request.session.get('bulto', {'productos': [], 'nombre': None})

        if request.method == 'POST' and form.is_valid():
            codigo_barras = form.cleaned_data['codigo_barras']

            try:
                # Buscar el producto en la base de datos
                producto = Producto.objects.get(codigo_barras=codigo_barras)

                encontrado = False
                for p in bulto['productos']:
                    if p['codigo_de_barras'] == producto.codigo_barras:
                        p['cantidad'] += 1
                        encontrado = True
                        break

                if not encontrado:
                    bulto['productos'].append({
                        'codigo_de_barras': producto.codigo_barras,
                        'nombre': producto.codigo,
                        'cantidad': 1
                    })

                nombre = form.cleaned_data['nombre_bulto']
                bulto['nombre'] = nombre
        
                request.session['bulto'] = bulto
                messages.success(request, 'Producto agregado correctamente.')
                form.fields['codigo_barras'].initial = ''

            except Producto.DoesNotExist:
                messages.error(request, 'El producto no fue encontrado.')

            return render(request, 'bultos/bulto_forms.html', {'form': form, 'productos': bulto['productos']})

        return render(request, 'bultos/bulto_forms.html', {'form': form, 'productos': bulto['productos']})
    except Exception:
        return server_error(request)
############################################################################
@login_required(login_url='login_app:login')
def crear_bulto(request):
    try:
        bulto = request.session.get('bulto', {'productos': []})

        if bulto['productos']:
            nombre_bulto = bulto.get('nombre')

            nuevo_bulto = Bulto(nombre=nombre_bulto, productos=bulto['productos'])
            nuevo_bulto.save()

            del request.session['bulto']

            registrarLoggerUsuario(request.user, "El usuario insertó un bulto en la base de datos")

            add_bootstrap_message(request, 'success', '¡Bulto creado exitosamente!')

            return redirect('almacen_app:bultos')
        else:
            messages.error(request, 'No hay productos en el bulto.')

        return redirect('almacen_app:bultos')
    except Exception:
        return server_error(request)
#############################################################################
@login_required(login_url='login_app:login')
def actualizar_cantidad_productos(request, bulto_id):
    bulto = get_object_or_404(Bulto, id=bulto_id)
    try:
        for producto_data in bulto.productos:
            codigo_barras = producto_data['codigo_de_barras']
            cantidad = producto_data['cantidad']
        
            producto = get_object_or_404(Producto, codigo_barras=codigo_barras)
            producto.cantidad += cantidad
            producto.save()
    
        bulto.insertado = True
        bulto.save()
        add_bootstrap_message(request, 'success', '¡Bulto entrado exitosamente!')
        # Obtener el contexto necesario para la renderización
        bultos = Bulto.objects.all()
        context = {'bultos': bultos}
    
        return render(request, 'bultos/bultos.html', context)
    except Exception:
        return server_error(request) 
#############################################################################
@login_required(login_url='login_app:login')
def mostrar_bultos(request):
    try:
        bultos = Bulto.objects.all()

        # Comprobar si la variable de sesión está llena
        if 'bulto' in request.session:
            bulto = request.session['bulto']
            if bulto['productos']:
                # Vaciar la variable de sesión
                del request.session['bulto']
        return render(request, 'bultos/bultos.html', {'bultos': bultos})
    except Exception:
        return server_error(request)
#############################################################################
@login_required(login_url='login_app:login')
def detalles_bulto(request, bulto_id):
    bulto = Bulto.objects.get(id=bulto_id)
    return render(request,"bultos/bulto_view.html", {"bulto": bulto})
##############################################################################
@login_required(login_url='login_app:login')
def delete_bulto(request, bulto_id):
    try:
        bult = get_object_or_404(Bulto, id=bulto_id)

        if request.method == 'POST':
            bult.delete()
            registrarLoggerUsuario(request.user, "El usuario eliminó un bulto en la base de datos")
            add_bootstrap_message(request, 'success', '¡Bulto eliminado exitosamente!')
            return redirect(reverse('almacen_app:bultos'))  # Redirige a la lista de productos o a otra vista

        context = {
            'bultos': bult,
        }
        return render(request, 'confirm_delete.html', context)
    except Exception:
        return server_error(request)
###########################################################################
@login_required(login_url='login_app:login')
def buscar_producto(request):
    try:
        if request.method == 'POST':
            codigo_barras = request.POST.get('codigo_barras')
            try:
                producto = Producto.objects.get(codigo_barras=codigo_barras)
                response_data = {
                    'success': True,
                    'codigoBarras': producto.codigo_barras,
                    'nombre': producto.nombre
                }
            except Producto.DoesNotExist:
                response_data = {'success': False}

            return JsonResponse(response_data)

        return JsonResponse({'success': False})
    except Exception:
        return server_error(request)

##############################################################################3
@login_required(login_url='login_app:login')
def detalle_producto(request, producto_id):
    try:
        producto = get_object_or_404(Producto, pk=producto_id)
        estantes, almacenes = get_nombres_estantes_y_almacenes(producto)

        # Descripción formateada
        descripcion_formateada = []
        nivel_actual = None
        if producto.descripcion:
            for line in producto.descripcion.split('\n'):
                if line.startswith('*'):
                    if nivel_actual:
                        descripcion_formateada.append(nivel_actual)
                    nivel_actual = {
                        'nivel': line[1:].strip(),
                        'subniveles': []
                    }
                elif nivel_actual:
                    subniveles = [subnivel.strip() for subnivel in line.split(';') if subnivel.strip()]  # Solo agregar subniveles no vacíos
                    nivel_actual['subniveles'].extend(subniveles)

            if nivel_actual:
                descripcion_formateada.append(nivel_actual)
        else:
            descripcion_formateada.append({
                'nivel': '-',
                'subniveles': []
            })

        # OEM formateada
        oem_formateada = []
        nivel_actual = None
        if producto.oem:
            for line in producto.oem.split('\n'):
                if line.startswith('*'):
                    if nivel_actual:
                        oem_formateada.append(nivel_actual)
                    nivel_actual = {
                        'nivel': line[1:].strip(),
                        'subniveles': []
                    }
                elif nivel_actual:
                    subniveles = [subnivel.strip() for subnivel in line.split(';') if subnivel.strip()]  # Solo agregar subniveles no vacíos
                    nivel_actual['subniveles'].extend(subniveles)

            if nivel_actual:
                oem_formateada.append(nivel_actual)
        else:
            oem_formateada.append({
                'nivel': '-',
                'subniveles': []
            })

        # Compatibilidad formateada
        compatibilidad_formateada = []
        if producto.compatibilidad:
            compatibilidad = producto.compatibilidad.split('\n')
            for compatibilidad_word in compatibilidad:
                compatibilidad_formateada.append({
                    'nivel': compatibilidad_word.strip(),
                    'subniveles': []
                })
        else:
            compatibilidad_formateada.append({
                'nivel': '-',
                'subniveles': []
            })

        return render(request, 'producto-view.html', {
            'producto': producto,
            'descripcion_formateada': descripcion_formateada,
            'compatibilidad_formateada': compatibilidad_formateada,
            'almacenes_con_producto': almacenes,
            'estantes_con_producto': estantes,
            'oem_formateada': oem_formateada
        })
    except Producto.DoesNotExist:
        raise Http404("Producto no existe")
    except Exception as e:
        return render(request, 'error.html', {'error_message': str(e)})

###########################################################################
##########################LECTOR-BARRAS-DESPACHO###########################
@login_required(login_url='login_app:login')
def escanear_codigo_barras_despacho(request):
    try:
        form = EscanearCodigoDeBarrasDespachoForm(request.POST or None)
        despacho = request.session.get('despacho', {'productos': [], 'nombre': None, 'cliente': None})

        if request.method == 'POST' and form.is_valid():
            codigo_barras = form.cleaned_data['codigo_barras']

            try:
                # Buscar el producto en la base de datos
                producto = Producto.objects.filter(Q(codigo_barras=codigo_barras) | Q(codigo=codigo_barras)).first()

                encontrado = False
                for p in despacho['productos']:
                    if p['codigo_de_barras'] == producto.codigo_barras or p['nombre'] == producto.codigo:
                        p['cantidad'] += 1
                        encontrado = True
                        break

                if not encontrado:
                    despacho['productos'].append({
                        'codigo_de_barras': producto.codigo_barras,
                        'nombre': producto.codigo,
                        'cantidad': 1
                    })

            
                nombre = form.cleaned_data['nombre_despacho']
                despacho['nombre'] = nombre
            
                nombre_cliente = form.cleaned_data['nombre_cliente']
                despacho['cliente'] = nombre_cliente
        
                request.session['despacho'] = despacho
                add_bootstrap_message(request, 'success', '¡Producto agregado exitosamente!')
            
                form.fields['codigo_barras'].initial = ''

            except Producto.DoesNotExist:
                add_bootstrap_message(request, 'warning', '¡El producto no fue encontrado!')

            return render(request, 'despachos/despacho_forms.html', {'form': form, 'productos': despacho['productos']})

        return render(request, 'despachos/despacho_forms.html', {'form': form, 'productos': despacho['productos']})
    except Exception:
        return server_error(request)

########################################################################
@login_required(login_url='login_app:login')
def crear_despacho(request):
    try:
        despacho = request.session.get('despacho', {'productos': []})

        if despacho['productos']:
            nombre_despacho = despacho.get('nombre')
            nombre_cliente = despacho.get('cliente')

            productos_despacho = despacho['productos']

            productos_base_datos = Producto.objects.filter(
                codigo_barras__in=[producto['codigo_de_barras'] for producto in productos_despacho]
            )
        
            # Verificar que la cantidad en el despacho no sea mayor que la cantidad en la base de datos
            for producto_despacho in productos_despacho:
                producto_base_datos = productos_base_datos.get(codigo_barras=producto_despacho['codigo_de_barras'])
                if producto_despacho['cantidad'] > producto_base_datos.cantidad:
                    # La cantidad en el despacho es mayor que la cantidad disponible en la base de datos
                    add_bootstrap_message(request, 'warning', f"No se puede crear el despacho. La cantidad de {producto_base_datos.nombre} es insuficiente.")
                    return redirect('almacen_app:despachos')
            else:
                nuevo_despacho = Despacho(nombre=nombre_despacho, productos=despacho['productos'], cliente=nombre_cliente)
                nuevo_despacho.save()

                del request.session['despacho']

                registrarLoggerUsuario(request.user, "El usuario insertó un despacho en la base de datos")

                add_bootstrap_message(request, 'success', '¡Despacho creado exitosamente!')
                
                return redirect('almacen_app:despachos')
        else:
            add_bootstrap_message(request, 'warning', 'Por favor, verifique que los campos estén correctos.')

        return redirect('almacen_app:create_despacho')
    except Exception:
        return server_error(request)
#############################################################################
@login_required(login_url='login_app:login')
def mostrar_excel_despacho(request):
    try:
        if request.method == 'POST':
            form = DespachoForm(request.POST, request.FILES)
            if form.is_valid():
                despacho = form.save(commit=False)

                # Obtenemos el nombre y cliente del formulario
                nombre = form.cleaned_data['nombre']
                cliente = form.cleaned_data['cliente']
            
                # Agregamos el nombre, cliente y la fecha actual al despacho
                despacho.nombre = nombre
                despacho.cliente = cliente
                despacho.fecha = date.today()

                # Procesar el archivo Excel y crear los productos correspondientes
                excel_file = request.FILES['archivo_excel']
                wb = load_workbook(excel_file, read_only=True)
                ws = wb.active
                excel_preview_data = []

                for row in ws.iter_rows(min_row=2, values_only=True):
                    nombre_producto = row[0]
                    cantidad = row[1]

                    producto_data = {
                        'nombre': nombre_producto,
                        'cantidad': cantidad,
                    }
                    excel_preview_data.append(producto_data)

                # Guardamos los productos como lista JSON en el campo 'productos' del despacho
                productos_base_datos = Producto.objects.filter(
                codigo__in=[producto['nombre'] for producto in excel_preview_data]
            )
        
                # Verificar que la cantidad en el despacho no sea mayor que la cantidad en la base de datos
                for producto_despacho in excel_preview_data:
                    producto_base_datos = productos_base_datos.get(codigo=producto_despacho['nombre'])
                    if producto_despacho['cantidad'] > producto_base_datos.cantidad:
                        # La cantidad en el despacho es mayor que la cantidad disponible en la base de datos
                        add_bootstrap_message(request, 'warning', f"No se puede crear el despacho. La cantidad de {producto_base_datos.nombre} es insuficiente.")
                        return redirect('almacen_app:despachos')
                else:
                    despacho.productos = excel_preview_data
                    # Guardamos el despacho en la base de datos
                    despacho.save()
                    # decrementar_cantidad_productos(request, despacho.id)
                    registrarLoggerUsuario(request.user, "El usuario insertó un despacho desde un excel en la base de datos")
                    add_bootstrap_message(request, 'success', '¡Despacho insertado exitosamente!')
                    return redirect('almacen_app:despachos')  # Redirigir a una página de éxito o donde desees
        
            else:
                add_bootstrap_message(request, 'warning', 'Por favor. Verifique que los campos estén correctos')

        else:
            form = DespachoForm()

        return render(request, 'despachos/despacho_loadexcel.html', {'form': form})
    except Exception:
        return server_error(request)
###############################################################################

#############################################################################
@login_required(login_url='login_app:login')
def decrementar_cantidad_productos(request, despacho_id):
    try:
        despacho = get_object_or_404(Despacho, id=despacho_id)
    
        for producto_data in despacho.productos:
            codigo_barras = producto_data['codigo_de_barras']
            cantidad = producto_data['cantidad']
        
            producto = get_object_or_404(Producto, codigo_barras=codigo_barras)
            producto.cantidad -= cantidad
            producto.save()
    
        despacho.insertado = True
        despacho.save()
        add_bootstrap_message(request, 'success', '¡Despachado exitosamente!')
        # Obtener el contexto necesario para la renderización
        despachos = Despacho.objects.all()
        context = {'despachos': despachos}
    
        return render(request, 'despachos/despachos.html', context)
    except Exception:
        return server_error(request)
#############################################################################
@login_required(login_url='login_app:login')
def mostrar_despachos(request):
    despachos = Despacho.objects.all()

    # Comprobar si la variable de sesión está llena
    if 'despacho' in request.session:
        despacho = request.session['despacho']
        if despacho['productos']:
            # Vaciar la variable de sesión
            del request.session['despacho']
    return render(request, 'despachos/despachos.html', {'despachos': despachos})
#############################################################################
@login_required(login_url='login_app:login')
def detalles_despacho(request, despacho_id):
    despacho = Despacho.objects.get(id=despacho_id)
    return render(request,"despachos/despacho_views.html", {"despacho": despacho})
##############################################################################
@login_required(login_url='login_app:login')
def delete_despacho(request, despacho_id):
    try:
        despacho = get_object_or_404(Despacho, id=despacho_id)

        if request.method == 'POST':
            despacho.delete()
            registrarLoggerUsuario(request.user, "El usuario eliminó un despacho en la base de datos")
            add_bootstrap_message(request, 'success', '¡Despacho eliminado exitosamente!')
            return redirect(reverse('almacen_app:despachos'))  # Redirige a la lista de productos o a otra vista

        context = {
            'despachos': despacho,
        }
        return render(request, 'confirm_delete.html', context)
    except Exception:
        return server_error(request)
###########################################################################

@login_required(login_url='login_app:login')
def listar_almacenes(request):
    almacenes = Almacen.objects.all()
    context = {'almacenes': almacenes}
    return render(request, 'almacen/almacen.html', context)


@login_required(login_url='login_app:login')
def crear_almacen(request):
    try:
        if request.method == 'POST':
            form = AlmacenForm(request.POST)
            if form.is_valid():
                nombre = form.cleaned_data['nombre']
                productos = form.cleaned_data['productos']
                estantes = form.cleaned_data['estantes']
            
                almacen = Almacen(nombre=nombre)
                almacen.save()
            
                if productos:
                    almacen.productos.set(productos)
            
                if estantes:
                    almacen.estantes.set(estantes)

                registrarLoggerUsuario(request.user, "El usuario insertó un almacén en la base de datos")
            
                add_bootstrap_message(request, 'success', '¡Almacén creado exitosamente!')
                return redirect(reverse('almacen_app:almacenes'))
            else:
                add_bootstrap_message(request, 'warning', 'Por favor. Verifique que los campos estén correctos')
        else:
            form = AlmacenForm()

        context = {
            'form': form,
            'is_editing': False,
        }
        return render(request, 'almacen/almacen_forms.html', context)
    except Exception:
        return server_error(request)
##############################################################################
@login_required(login_url='login_app:login')
def editar_almacen(request, almacen_id):
    try:
        almacen = get_object_or_404(Almacen, id=almacen_id)
        if request.method == 'POST':
            form = AlmacenForm(request.POST)
            if form.is_valid():
                nombre = form.cleaned_data['nombre']
                productos = form.cleaned_data['productos']
                estantes = form.cleaned_data['estantes']
            
                almacen.nombre = nombre
                almacen.save()
            
                almacen.productos.clear()
                if productos:
                    almacen.productos.set(productos)
            
                almacen.estantes.clear()
                if estantes:
                    almacen.estantes.set(estantes)
            
                registrarLoggerUsuario(request.user, "El usuario editó un almacén en la base de datos")
                add_bootstrap_message(request, 'success', '¡Almacén actualizado exitosamente!')
                return redirect(reverse('almacen_app:almacenes'))
            else:
                add_bootstrap_message(request, 'warning', 'Por favor. Verifique que los campos sean correctos')
        else:
            form = AlmacenForm(instance=almacen)

        context = {
            'form': form,
            'is_editing': True,
        }
        return render(request, 'almacen/almacen_forms.html', context)
    except Exception:
        return server_error(request)
#######################################################################################
@login_required(login_url='login_app:login')
def eliminar_almacen(request, almacen_id):
    try:
        almacen = get_object_or_404(Almacen, id=almacen_id)
        almacen.delete()
        registrarLoggerUsuario(request.user, "El usuario eliminó un almacén en la base de datos")
        add_bootstrap_message(request, 'success', '¡Almacén eliminado exitosamente!')
        return redirect(reverse('almacen_app:almacenes'))
    except Exception:
        return server_error(request)
#########################################################################
@login_required(login_url='login_app:login')
def detalle_almacen(request, almacen_id):
    almacen = get_object_or_404(Almacen, pk=almacen_id)
    return render(request, 'almacen/almacen_views.html', {'almacen': almacen})
###########################################################################
@login_required(login_url='login_app:login')
def listar_estantes(request):
    estantes = Estante.objects.all()
    context = {'estantes': estantes}
    return render(request, 'estante/estante.html', context)
############################################################################
@login_required(login_url='login_app:login')
def detalle_estante(request, estante_id):
    estante = get_object_or_404(Estante, pk=estante_id)
    return render(request, 'estante/estante_views.html', {'estante': estante})
###############################################################################
@login_required(login_url='login_app:login')
def crear_estante(request):
    try:
        if request.method == 'POST':
            form = EstanteForm(request.POST)
            if form.is_valid():
                nombre = form.cleaned_data['nombre']
                productos = form.cleaned_data['productos']
            
                estante = Estante(nombre=nombre)
                estante.save()
            
                if productos:
                    estante.productos.set(productos)
            
                registrarLoggerUsuario(request.user, "El usuario insertó un estante en la base de datos")
                add_bootstrap_message(request, 'success', '¡Estante creado exitosamente!')
                return redirect(reverse('almacen_app:almacenes'))
            else:
                add_bootstrap_message(request, 'warning', 'Por favor. Verifique que los campos sean correctos')
        else:
            form = EstanteForm()

        context = {
            'form': form,
            'is_editing': False,
        }
        return render(request, 'estante/estante_forms.html', context)
    except Exception:
        return server_error(request)
########################################################################
@login_required(login_url='login_app:login')
def editar_estante(request, estante_id):
    try:
        estante = get_object_or_404(Estante, id=estante_id)
        if request.method == 'POST':
            form = EstanteForm(request.POST)
            if form.is_valid():
                nombre = form.cleaned_data['nombre']
                productos = form.cleaned_data['productos']
            
                estante.nombre = nombre
                estante.save()
            
                estante.productos.clear()
                if productos:
                    estante.productos.set(productos)
            
                registrarLoggerUsuario(request.user, "El usuario editó un estante en la base de datos")
                add_bootstrap_message(request, 'success', '¡Estante actualizado exitosamente!')
                return redirect(reverse('almacen_app:almacenes'))
            else:
                add_bootstrap_message(request, 'Por favor. Verifique que los campos sean correctos')
        else:
            form = EstanteForm(instance=estante)

        context = {
            'form': form,
            'is_editing': True,
        }
        return render(request, 'estante/estante_forms.html', context)
    except Exception:
        return server_error(request)
############################################################################
@login_required(login_url='login_app:login')
def eliminar_estante(request, estante_id):
    try:
        estante = get_object_or_404(Estante, id=estante_id)
        estante.delete()
        registrarLoggerUsuario(request.user, "El usuario eliminó un estante en la base de datos")
        add_bootstrap_message(request, 'success', '¡Estante eliminado exitosamente!')
        return redirect(reverse('almacen_app:almacenes'))
    except Exception:
        return server_error(request)
###############################EXPORT-EXCEL################################
@login_required(login_url='login_app:login')
def export_excel(request):
    try:
        objeto_a_trabajar = Producto.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.append(['Nombre', 'Código de Barras', 'Producto', 'Descripción', 'Precio', 'Cantidad', 'Compatibilidad', 'Clasificación'])

        contador = 2
        for q in objeto_a_trabajar:
            column = str(contador) 
            ws['A' + column] = q.codigo
            ws['B'+ column] = q.codigo_barras
            ws['C'+ column] = q.nombre
            ws['D'+ column] = q.descripcion
            ws['E'+ column] = q.precio
            ws['F'+ column] = q.cantidad
            ws['G'+ column] = q.compatibilidad
            ws['H'+ column] = q.campo_predefinido
            contador += 1

        # Establecer estilo para borde de celda
        border = Border(left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000'))

        # Establecer estilo para alineación de texto
        alignment = Alignment(horizontal="center", vertical="center")

        # Establecer estilo para fuente
        font = Font(name='Calibri', size=11, bold=True)

        # Agregar fila de título
        ws.insert_rows(1)
        ws['A1'] = "Listado de Productos"

        # Establecer estilo para título
        title_cell = ws['A1']
        title_cell.font = Font(name='Calibri', size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Combinar celdas de título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

        # Aplicar estilo a la fila de encabezado
        for cell in ws[2]:
            cell.border = border
            cell.alignment = alignment
            cell.font = font

        # Establecer ancho de columna
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 12

        # Establecer título de la hoja de cálculo
        ws.title = "Listado de Productos"

        # Establecer estilo para borde de tabla
        border = Border(left=Side(style='medium', color='000000'),
                        right=Side(style='medium', color='000000'),
                        top=Side(style='medium', color='000000'),
                        bottom=Side(style='medium', color='000000'))

        for row in ws.iter_rows(min_row=2, max_col=8, max_row=objeto_a_trabajar.count() + 1):
            for cell in row:
                cell.border = border

        # Establecer estilo para alineación de texto y fuente
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        font = Font(name='Calibri', size=10)

        for row in ws.iter_rows(min_row=2, max_col=8, max_row=objeto_a_trabajar.count() + 1):
            for cell in row:
                cell.alignment = alignment
                cell.font = font

        # Establecer estilo para borde de la última fila
        border_last_row = Border(left=Side(style='medium', color='000000'),
                                right=Side(style='medium', color='000000'),
                                bottom=Side(style='medium', color='000000'))

        # Establecer estilo para alineación de texto y fuente de la última fila
        alignment_last_row = Alignment(horizontal="center", vertical="center", wrap_text=True)
        font_last_row = Font(name='Calibri',size=10)

        # Aplicar estilo a la última fila
        for cell in ws[objeto_a_trabajar.count() + 2]:
            cell.border = border_last_row
            cell.alignment = alignment_last_row
            cell.font = font_last_row

        # Centrar la tabla horizontalmente
        ws.center_horizontally = True

        # Centrar el título verticalmente en la página
        ws.page_setup.center_header = "Listado de Productos"

        nombre_archivo = "Listado de Productos.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        registrarLoggerUsuario(request.user, "El usuario exportó listado de productos de la base de datos")
        return response
    except Exception:
        return server_error(request)
###############################IMPORTAR-EXCEL#############################
@login_required(login_url='login_app:login')
def cargar_productos_excel(request):
    try:
        if request.method == 'POST':
            form = CargarProductosForm(request.POST, request.FILES)
            if form.is_valid():
                archivo_excel = request.FILES['archivo_excel']
                wb = load_workbook(filename=archivo_excel, read_only=True)
                ws = wb.active

                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Ignorar filas vacías
                    if not any(row):
                        continue
                    
                    # Asignar valores predeterminados si los campos son nulos
                    row = [col if col is not None else '-' for col in row]
                    
                    # Obtener o crear el producto
                    codigo_excel = row[0] if row[0] != '-' else None
                    producto_existente = Producto.objects.filter(codigo=codigo_excel).first() if codigo_excel else None

                    if producto_existente:
                        producto_existente.cantidad += row[6] if isinstance(row[6], (int, float)) else 0
                        producto_existente.save()
                    else:
                        nuevo_producto = Producto(
                            codigo=row[0] if row[0] != '-' else None,
                            codigo_barras=row[1] if row[1] != '-' else None,
                            nombre=row[2] if row[2] != '-' else None,
                            descripcion=row[3] if row[3] != '-' else None,
                            oem=row[4] if row[4] != '-' else None,
                            precio=row[5] if isinstance(row[5], (int, float)) else 0,
                            cantidad=row[6] if isinstance(row[6], (int, float)) else 0,
                            compatibilidad=row[7] if row[7] != '-' else None,
                            campo_predefinido=row[8] if row[8] != '-' else None,
                        )
                        nuevo_producto.save()

                registrarLoggerUsuario(request.user, "El usuario insertó los productos desde un excel en la base de datos")
                add_bootstrap_message(request, 'success', '¡Excel importado exitosamente!')
                return redirect(reverse('login_app:index'))
            else:
                add_bootstrap_message(request, 'warning', 'Ha ocurrido un error en la importación')
        else:
            form = CargarProductosForm()
        return render(request, 'product_excel_forms.html', {'form': form})
    except Exception as e:
        # Registrar el error y mostrar un mensaje de error al usuario
        logger.error(f"Error importing products: {e}")
        add_bootstrap_message(request, 'warning', 'Ha ocurrido un error en la importación. Por favor, intente nuevamente.')
        return redirect(reverse('login_app:index'))
##########################EXPORTAR-EXCEL-DESPACHO##########################
@login_required(login_url='login_app:login')
def exportar_despacho_excel(request, despacho_id):
    try:
        try:
            despacho = Despacho.objects.get(id=despacho_id)
        except Despacho.DoesNotExist:
            return HttpResponse("Despacho no encontrado")

        # Crear un nuevo libro de Excel
        wb = Workbook()

        # Seleccionar la hoja activa
        ws = wb.active

        # Establecer el nombre de la hoja
        ws.title = f"{despacho.nombre}"

        # Establecer el ancho de las columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20

        # Establecer el nombre del despacho en la primera fila
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
        cell = ws.cell(row=1, column=2)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.value = despacho.nombre

        # Establecer el nombre del cliente y la fecha en la segunda fila
        ws.cell(row=2, column=1).value = f"Cliente: {despacho.cliente}"
        ws.cell(row=2, column=5).value = f"Fecha: {despacho.fecha.strftime('%d/%m/%Y')}"

        # Agregar la tabla de productos en la tercera fila
        ws.cell(row=4, column=2).value = "Producto"
        ws.cell(row=4, column=3).value = "Codigo de barras"
        ws.cell(row=4, column=4).value = "Cantidad"
        if despacho.productos:
            for i, producto in enumerate(despacho.productos):
                ws.cell(row=i+5, column=2).value = producto['nombre']
                ws.cell(row=i+5, column=3).value = producto['codigo_de_barras']
                ws.cell(row=i+5, column=4).value = producto['cantidad']

        # Establecer el estilo de la tabla
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=2, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.row == 4:
                    # Establecer el borde superior de la tabla
                    cell.border = Border(left=Side(border_style='thin'),
                                        right=Side(border_style='thin'),
                                        top=Side(border_style='thin'),
                                        bottom=Side(border_style=None))
                elif cell.row == ws.max_row:
                    # Establecer el borde inferior de la tabla
                    cell.border = Border(left=Side(border_style='thin'),
                                        right=Side(border_style='thin'),
                                        top=Side(border_style=None),
                                        bottom=Side(border_style='thin'))
                elif cell.column in (2, 3, 4):
                    # Eliminar los bordes interiores de la tabla
                    cell.border = Border(left=Side(border_style=None),
                                        right=Side(border_style=None),
                                        top=Side(border_style=None),
                                        bottom=Side(border_style=None))

        # Establecer el borde exterior de todo el contenido del Excel
        for row in ws.iter_rows():
            for cell in row:
                if cell.row == 1 or cell.column == 1 or cell.row == ws.max_row or cell.column == ws.max_column:
                    cell.border = Border(left=Side(border_style='medium'),
                                        right=Side(border_style='medium'),
                                        top=Side(border_style='medium'),
                                        bottom=Side(border_style='medium'))

        # Agregar una línea en blanco
        ws.cell(row=ws.max_row+1, column=1)

        # Agregar la raya y las firmas
        ws.cell(row=ws.max_row+4, column=1).border = Border(top=Side(border_style='thin'))
        ws.cell(row=ws.max_row, column=1).value = "Firma del Trabajador"
        ws.cell(row=ws.max_row, column=5).border = Border(top=Side(border_style='thin'))
        ws.cell(row=ws.max_row, column=5).value = "Firma del Cliente"

        # Establecer el estilo de las firmas
        for cell in ws.iter_rows(min_row=ws.max_row-1, max_row=ws.max_row, min_col=1, max_col=5):
            for c in cell:
                c.alignment= Alignment(horizontal='center', vertical='center')
                c.font = c.font.copy(bold=True)

        # Configurar la respuesta para descargar el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename= Despacho {despacho.cliente} {despacho.fecha} .xlsx'
        wb.save(response)
        registrarLoggerUsuario(request.user, "El usuario exportó un despacho a excel")

        return response
    except Exception:
        return server_error(request)
###############################FUNCIONES-INDEX#############################
@login_required(login_url='login_app:login')
def calcular_sumatoria_precios(request):
    try:
        # Obtener todos los precios como cadenas
        precios_str = Producto.objects.filter(cantidad__gt=0).values_list('precio', flat=True)
        
        # Convertir los precios a números
        precios_numeros = [float(precio.replace(',', '.')) for precio in precios_str]
        
        # Sumar los precios
        sumatoria = sum(precios_numeros)
        
        # Devolver la sumatoria como respuesta JSON
        return int(sumatoria)
    except Exception:
        return server_error(request)

@login_required(login_url='login_app:login')
def calcular_cantidad_productos(request):
    try:
        cantidad_productos = Producto.objects.count()
        return cantidad_productos
    except Exception:
        return server_error(request)

@login_required(login_url='login_app:login')
def calcular_total_productos_agotados(request):
    try:
        total_agotados = Producto.objects.filter(cantidad=0).count()
        return total_agotados
    except Exception:
        return server_error(request)

@login_required(login_url='login_app:login')
def calcular_porcentaje_disponibilidad(request):
    try:
        total_productos = Producto.objects.count()
        total_disponibles = Producto.objects.filter(cantidad__gt=0).count()

        if total_productos > 0:
            porcentaje = (total_disponibles / total_productos) * 100
        else:
            porcentaje = 0

        return int(porcentaje)
    except Exception:
        return server_error(request)

@login_required(login_url='login_app:login')
def obtener_datos_area(request):
    try:
        tipos_productos = Producto.objects.values_list('campo_predefinido', flat=True).distinct()
        sumas_precios = []
    
        for tipo_producto in tipos_productos:
            sumatoria = Producto.objects.filter(campo_predefinido=tipo_producto).aggregate(total=Sum('precio'))['total']
            if sumatoria is None:
                sumatoria = 0
            sumas_precios.append(sumatoria)
    
        return tipos_productos, sumas_precios
    except Exception:
        return server_error(request)

logger = logging.getLogger(__name__)

###############################-BUSCAR-PRODUCTO-GENERAL-###################################
def get_nombres_estantes(producto):
    try:
        estantes = Estante.objects.filter(productos=producto)
        if estantes:
            return [estante.nombre for estante in estantes]
        else:
            return ["Todavía no está ubicado en ningún estante"]
    except Exception:
        return server_error(producto)

def get_nombres_almacenes(producto):
    try:
        almacenes = Almacen.objects.filter(productos=producto)

        if almacenes.exists():
            return [almacen.nombre for almacen in almacenes]
        else:
            return ["Todavía no está ubicado en ningún almacén en específico"]
    except Exception:
        return server_error(producto)
    
def get_nombres_estantes_y_almacenes(producto):
    try:
        estantes = Estante.objects.filter(productos=producto)
        almacenes_por_producto = Almacen.objects.filter(productos=producto)
        almacenes_por_estante = Almacen.objects.filter(estantes__in=estantes)

        nombres_estantes = [estante.nombre for estante in estantes] if estantes.exists() else ["Todavía no está ubicado en ningún estante en específico"]
        nombres_almacenes = []

        if almacenes_por_producto.exists():
            nombres_almacenes += [almacen.nombre for almacen in almacenes_por_producto]

        if almacenes_por_estante.exists():
            nombres_almacenes += [almacen.nombre for almacen in almacenes_por_estante]

        nombres_almacenes = nombres_almacenes if nombres_almacenes else ["Todavía no está ubicado en ningún almacén en específico"]

        return nombres_estantes, nombres_almacenes
    except Exception:
        return server_error(producto)

def buscar_producto_por_codigo_de_barras(request, busqueda):
    logger.info("Solicitud recibida para buscar producto")

    try:
        if not busqueda:
            return JsonResponse({'error': 'Ingrese un valor para buscar'})

        # Realiza la búsqueda por código de barras, código o nombre
        producto = get_object_or_404(Producto, Q(codigo_barras=busqueda) | Q(codigo=busqueda) | Q(nombre__icontains=busqueda) | Q(oem__icontains=busqueda))

        estantes, almacenes = get_nombres_estantes_y_almacenes(producto)

        data = {
            'nombre': producto.codigo,
            'producto': producto.nombre,
            'codigo_barra': producto.codigo_barras,
            'descripcion': producto.descripcion,
            'oem': producto.oem,
            'precio': str(producto.precio),
            'cantidad': producto.cantidad,
            'compatibilidad': producto.compatibilidad,
            'ubicacion': producto.campo_predefinido,
            'estantes': estantes,
            'almacenes': almacenes,
            'imagen_url': producto.imagen.url if producto.imagen else '/media/productos/tmp_img.png',
        }

        return JsonResponse(data)
    except Producto.DoesNotExist:
        return server_error(request)

###########################################################################
##############################EXPORT-EXCEL###################################
def descargar_excel_productos(request):
    try:
        archivo_origen = 'media/excel-export/productos.xlsx'
    
        nombre_descarga = 'productos.xlsx'
    
        response = FileResponse(open(archivo_origen, 'rb'))
        response['Content-Disposition'] = f'attachment; filename="{nombre_descarga}"'
    
        return response
    except Exception:
        return server_error(request)

def descargar_excel_despacho(request):
    try:
        archivo_origen = 'media/excel-export/Despacho_Importar.xlsx'
    
        nombre_descarga = 'Importar-Despacho.xlsx'
    
        response = FileResponse(open(archivo_origen, 'rb'))
        response['Content-Disposition'] = f'attachment; filename="{nombre_descarga}"'
    
        return response
    except Exception:
        return server_error(request)

###########################################################################
@login_required(login_url='login_app:login')
def get_agregar_producto_a_almacenes(request, producto_id):
    almacenes = Almacen.objects.all().order_by('nombre')
    return render(request, "almacen/add_almacen.html", 
                  { "almacenes" : almacenes, "producto_id": producto_id})
    
@login_required(login_url='login_app:login')
def agregar_producto_a_almacenes(request, producto_id):
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        
        almacenes_ids = request.POST.getlist('almacen_id')
        
        producto = Producto.objects.get(pk=product_id)
        
        for almacen_id in almacenes_ids:
            almacen = Almacen.objects.get(pk=almacen_id)
            
            almacen.productos.add(producto)
        
        return redirect('almacen_app:productos')
    else:
        almacenes = Almacen.objects.all().order_by('nombre')
        
        return render(request, "almacen/add_almacen.html", { "almacenes": almacenes, "producto_id": producto_id })

@login_required(login_url='login_app:login')
def get_agregar_producto_a_estantes(request, producto_id):
    estantes = Estante.objects.all().order_by('nombre')
    return render(request, "estante/add_estante.html", 
                  { "estantes" : estantes, "producto_id": producto_id})
    
@login_required(login_url='login_app:login')
def agregar_producto_a_estantes(request, producto_id):
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        
        estantes_ids = request.POST.getlist('estante_id')
        
        producto = Producto.objects.get(pk=product_id)
        
        for estante_id in estantes_ids:
            estante = Estante.objects.get(pk=estante_id)
            
            estante.productos.add(producto)
        
        return redirect('almacen_app:productos')
    else:
        estantes = Estante.objects.all().order_by('nombre')
        
        return render(request, "estante/add_estante.html", { "estantes": estantes, "producto_id": producto_id })
    

def filtrar_productos_por_valor(request, valor):
    productos_filtrados = Producto.objects.filter(
        Q(nombre__icontains=valor) |
        Q(descripcion__icontains=valor) |
        Q(codigo__icontains=valor) |
        Q(codigo_barras__icontains=valor) |
        Q(precio__icontains=valor) |
        Q(cantidad__icontains=valor) |
        Q(compatibilidad__icontains=valor) |
        Q(oem__icontains=valor)
    ).order_by('codigo')
    return render(request, "search_productos.html", 
                  { "productos" : productos_filtrados})
#################################################################################################


def add_bootstrap_message(request, level, message):
    tags = {
        'success': 'alert-success',
        'info': 'alert-info',
        'warning': 'alert-warning',
        'error': 'alert-danger',
    }
    messages.add_message(request, getattr(messages, level.upper()), message, extra_tags='alert {}'.format(tags[level]))